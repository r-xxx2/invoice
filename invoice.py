# To add a new cell, type '# %%'
# To add a new markdown cell, type '# %% [markdown]'
# %%
import pandas as pd
import openpyxl as px
import datetime
import calendar
import datetime
from dateutil.relativedelta import relativedelta

#----------------------
# 【１】 年月日を入力
#----------------------
invoice_month = input("請求書を作成する年月を入れてください(例：2110)")

#入力された値をdataに変換
dte = datetime.datetime.strptime(invoice_month, "%y%m")
print(dte)



#エラー処理(日付型に変換できないときに表示させたい) 
    # try:
    #     print(dte)
    # except unconverted data remains:
    #     print("正しい値を入力してください。")



#----------------------------------
# 【６】 日付、お振込み期限
#----------------------------------


#--------- 日本語でstrftimeが表示できないエラーの回避 -------------------------

import locale
locale.setlocale(locale.LC_CTYPE, "Japanese_Japan.932")

#----------------------------------------------------------------------------


#選択月の末日を習得
def EndOfMonth(dte):
    return dte.replace(day=calendar.monthrange(dte.year, dte.month)[1])

EOM = EndOfMonth(dte)
print(EOM.strftime("%Y年%m月%d日"))


#翌月末を取得
next_month = EOM + relativedelta(months=1)
# print(next_month)





#----------------------
# 【２】 請求書一覧から顧客CDをDFに取得
#----------------------
customer_df = pd.read_excel("請求一覧創造太郎.xlsx", sheet_name="顧客管理テーブル", index=0, header=1)
display(customer_df)




#----------------------
# 【３】 inputからその月のシートを取得
#----------------------

# invoice_month = input("請求書を作成する年月を入れてください(例：2110)")

sheetname = "請求一覧" + invoice_month

month_df = pd.read_excel("請求一覧創造太郎.xlsx", sheet_name=sheetname, header=2, index=False)
display(month_df)

month_df = month_df.merge(customer_df, left_on="相手先コード", right_on="顧客ＣＤ", how="left")
display(month_df)


# ４ 相手先コードのユニークを作成
sup_cd = month_df["相手先コード"].unique()
print(sup_cd)
sup_cd.sort()
print(sup_cd)


#-------------------------------------
# 【７】 各会社ごとにExcelに出力
#-------------------------------------


#----------- 出力セルの定義 ----------------

date_loc = "H1"                # 日付
ivno_loc = "H2"                # 請求書No.
Co_loc = "A6"                  # 会社名
sum_loc = "G28"                # 合計金額
dl_loc = "B35"                 # 振込期限

#------------------------------------------





###################
#メインのコード
###################


x = 1
for customer_cd in sup_cd:
    string = str(x).zfill(2) # 請求書No.用の連番
    
    wb = px.load_workbook("請求書創造太郎.xlsx")
    ws = wb["請求書元"]

    
    df = month_df[month_df["相手先コード"] == customer_cd]
    print(display(df))
    print("---------------")
    df_row = int(len(df))
    df_row += 1
    print(df_row)
    print(10 + 2)
    #------- 月一覧のDFからカラムごとにリスト化 --------------

    odr_no = df["受注No."].to_list()
    cus_cd = df["相手先コード"].to_list()
    cus_name = df["顧客名"].to_list()
    product_name1 = df["商品名１"].to_list()
    product_name2 = df["商品名２"].to_list()
    en = df["金額（税抜き）"].to_list()
    bikou = df["備考"].to_list()

    #-------------------------------------------------------


    ws[date_loc] = EOM.strftime("%Y年%m月%d日")            # 日付
    ws[ivno_loc] = invoice_month + string                  # 請求書No.
    ws[Co_loc] = cus_name[0] + " 御中"                     # 会社名
    ws[dl_loc] = "お振込み期限　　：　　" + next_month.strftime("%Y年%m月%d日")   # 振込期限

    j = 0
    for i in range(18,27,2):
        if j == len(df):
            break
        ws.cell(row= i, column=1).value = odr_no[j]           # 受注No.
        ws.cell(row= i, column=2).value = product_name1[j]    # 商品名１
        ws.cell(row= i+1, column=2).value = product_name2[j]  # 商品名２
        ws.cell(row= i, column=5).value = 1                   # 数量
        ws.cell(row= i, column=6).value = "式"                # 式
        ws.cell(row= i, column=7).value = en[j]               # 金額
        ws.cell(row= i, column=8).value = bikou[j]            # 備考
        j += 1

    total = df["金額（税抜き）"].sum()
    tax = total * 0.1

    # 商品名蘭の書式設定
    ws.cell(row= 18+len(df)*2+1 , column=2).value = "以上に掛かる消費税"     # 以上に掛かる消費税
    ws.cell(row= 18+len(df)*2+2 , column=2).value = "～　以　下　余　白　～"  # ～以下余白～
    ws.cell(row= 18+len(df)*2 , column=7).value = tax                      # 消費税
    ws[sum_loc] = total + tax                                                   # 合計金額

    # 書き出しのファイル設定
    wb.save(EOM.strftime("%Y年%m月") + "分請求書_【" + cus_name[0] + "様】.xlsx")
    x += 1


